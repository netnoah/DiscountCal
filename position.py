# position.py
"""Position tracking: load from Excel, calculate captured basis returns."""

import logging
from datetime import date, datetime
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)


def load_positions(filepath: str) -> list[dict]:
    """Load positions from Excel file.

    Expected columns (Chinese):
        合约, 买入日期, 买入价格, 买入时基准价, 是否已卖出, 吃到贴水

    Returns list of position dicts. Returns empty list if file missing or empty.
    Each dict includes row_index matching the actual Excel row number (1-based).
    """
    if not Path(filepath).exists():
        return []

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        wb.close()
        return []

    positions = []
    for idx, row in enumerate(rows):
        excel_row = idx + 2  # 1-indexed; row 1 is header, first data row is 2
        contract, buy_date_str, buy_price, base_price, sold_str, captured = row
        if not contract or not buy_date_str or not buy_price:
            continue

        if isinstance(buy_date_str, datetime):
            buy_date = buy_date_str.date()
        elif isinstance(buy_date_str, date):
            buy_date = buy_date_str
        else:
            buy_date = date.fromisoformat(str(buy_date_str))

        positions.append({
            "row_index": excel_row,
            "contract": str(contract).strip().upper(),
            "buy_date": buy_date,
            "buy_price": float(buy_price),
            "base_price_at_buy": float(base_price) if base_price else 0.0,
            "sold": str(sold_str).strip().upper() == "Y",
            "captured_basis": float(captured) if captured is not None else None,
        })

    wb.close()
    return positions


def save_captured_basis(filepath: str, row_index: int, value: float) -> None:
    """Write captured_basis value back to Excel for a sold position.

    row_index is the 1-based Excel row number (matching the row_index
    in position dicts). Column 6 is the "吃到贴水" column.
    """
    if not Path(filepath).exists():
        return
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    ws.cell(row=row_index, column=6, value=value)
    wb.save(filepath)
    wb.close()


def calculate_position_return(
    position: dict,
    current_futures_price: float,
    current_base_price: float,
    reference_date: date | None = None,
) -> dict:
    """Calculate captured basis metrics for a single position."""
    if reference_date is None:
        reference_date = date.today()
    else:
        reference_date = date.fromordinal(reference_date.toordinal()) if isinstance(reference_date, datetime) else reference_date

    buy_price = position["buy_price"]
    base_at_buy = position["base_price_at_buy"]
    buy_date = position["buy_date"]

    initial_basis = base_at_buy - buy_price
    current_basis = current_base_price - current_futures_price
    captured = initial_basis - current_basis

    holding_days = (reference_date - buy_date).days

    if holding_days > 0:
        annualized_return = captured / buy_price * 365 / holding_days * 100
    else:
        annualized_return = None

    if initial_basis != 0:
        convergence_pct = captured / initial_basis * 100
    else:
        convergence_pct = None

    return {
        "contract": position["contract"],
        "buy_price": buy_price,
        "current_price": current_futures_price,
        "initial_basis": round(initial_basis, 2),
        "current_basis": round(current_basis, 2),
        "captured_basis": round(captured, 2),
        "convergence_pct": round(convergence_pct, 2) if convergence_pct is not None else None,
        "holding_days": holding_days,
        "annualized_return": round(annualized_return, 2) if annualized_return is not None else None,
        "pnl": round(current_futures_price - buy_price, 2),
        "sold": position["sold"],
    }


def build_position_summary(
    position_returns: list[dict], total_count: int
) -> str:
    """Format position returns as WeChat notification text."""
    if not position_returns:
        return ""

    lines = [
        "",
        "📊 我的持仓收益",
        "────────────",
    ]

    unsold_count = 0
    for r in position_returns:
        if r.get("sold"):
            continue
        unsold_count += 1
        lines.append(f"🔹 {r['contract']} | 买入 {r['buy_price']:.1f}")

        captured = r["captured_basis"]
        captured_str = f"{captured:.1f}" if captured is not None else "N/A"
        lines.append(f"  当前价: {r['current_price']:.1f} | 已吃贴水: {captured_str}")

        conv = r["convergence_pct"]
        conv_str = f"{conv:.1f}%" if conv is not None else "N/A"
        annual = r["annualized_return"]
        annual_str = f"{annual:.1f}%" if annual is not None else "N/A"
        pnl = r["pnl"]
        pnl_str = f"{pnl:.1f}" if pnl is not None else "N/A"
        lines.append(f"  收敛: {conv_str} | 年化: {annual_str} | 浮盈: {pnl_str}")

    lines.append("────────────")
    lines.append(f"📌 持仓数: {total_count} | 未平仓: {unsold_count}")

    return "\n".join(lines)


def compute_position_returns(
    positions: list[dict],
    price_map: dict[str, float],
    near_price: float,
) -> list[dict]:
    """Calculate returns for all positions that have a matching price.

    Returns results for:
    - All unsold positions with a price match
    - Sold positions with no captured_basis (for write-back)

    Skips sold positions that already have captured_basis frozen.
    """
    results = []
    for pos in positions:
        # Skip sold positions that already have a frozen value
        if pos["sold"] and pos["captured_basis"] is not None:
            continue

        current_price = price_map.get(pos["contract"])
        if current_price is None:
            continue

        result = calculate_position_return(
            position=pos,
            current_futures_price=float(current_price),
            current_base_price=near_price,
        )
        results.append(result)
    return results
