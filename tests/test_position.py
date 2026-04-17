# tests/test_position.py
import pytest
from datetime import date
from pathlib import Path

import openpyxl

from position import load_positions, save_captured_basis, calculate_position_return, build_position_summary, compute_position_returns


class TestLoadPositions:
    def _create_test_excel(self, tmp_path, rows):
        """Helper to create a test positions.xlsx with given rows."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["合约", "买入日期", "买入价格", "买入时基准价", "是否已卖出", "吃到贴水"])
        for row in rows:
            ws.append(row)
        filepath = tmp_path / "positions.xlsx"
        wb.save(filepath)
        return str(filepath)

    def test_loads_unsold_position(self, tmp_path):
        filepath = self._create_test_excel(tmp_path, [
            ["I2609", "2026-04-10", 650.0, 720.0, "N", None],
        ])
        positions = load_positions(filepath)
        assert len(positions) == 1
        pos = positions[0]
        assert pos["contract"] == "I2609"
        assert pos["buy_date"] == date(2026, 4, 10)
        assert pos["buy_price"] == 650.0
        assert pos["base_price_at_buy"] == 720.0
        assert pos["sold"] is False
        assert pos["captured_basis"] is None
        # row_index must match actual Excel row (row 2 = first data row)
        assert pos["row_index"] == 2

    def test_loads_sold_position_with_captured_basis(self, tmp_path):
        filepath = self._create_test_excel(tmp_path, [
            ["I2605", "2026-03-15", 680.0, 710.0, "Y", 25.0],
        ])
        positions = load_positions(filepath)
        assert len(positions) == 1
        assert positions[0]["sold"] is True
        assert positions[0]["captured_basis"] == 25.0

    def test_loads_multiple_positions(self, tmp_path):
        filepath = self._create_test_excel(tmp_path, [
            ["I2609", "2026-04-10", 650.0, 720.0, "N", None],
            ["I2612", "2026-04-12", 630.0, 720.0, "N", None],
        ])
        positions = load_positions(filepath)
        assert len(positions) == 2
        assert positions[0]["row_index"] == 2
        assert positions[1]["row_index"] == 3

    def test_returns_empty_list_when_file_missing(self, tmp_path):
        filepath = str(tmp_path / "nonexistent.xlsx")
        positions = load_positions(filepath)
        assert positions == []

    def test_returns_empty_list_when_file_empty(self, tmp_path):
        filepath = self._create_test_excel(tmp_path, [])
        positions = load_positions(filepath)
        assert positions == []

    def test_skips_rows_with_missing_fields(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["合约", "买入日期", "买入价格", "买入时基准价", "是否已卖出", "吃到贴水"])
        ws.append(["I2609", "2026-04-10", 650.0, 720.0, "N", None])
        ws.append(["", "", "", "", "", ""])  # empty row — should be skipped
        ws.append(["I2612", "2026-04-12", 630.0, 720.0, "N", None])
        filepath = tmp_path / "positions.xlsx"
        wb.save(filepath)
        positions = load_positions(str(filepath))
        assert len(positions) == 2
        # First valid row is Excel row 2
        assert positions[0]["row_index"] == 2
        # Second valid row is Excel row 4 (row 3 was empty/skipped)
        assert positions[1]["row_index"] == 4


class TestSaveCapturedBasis:
    def _create_test_excel(self, tmp_path, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["合约", "买入日期", "买入价格", "买入时基准价", "是否已卖出", "吃到贴水"])
        for row in rows:
            ws.append(row)
        filepath = tmp_path / "positions.xlsx"
        wb.save(filepath)
        return str(filepath)

    def test_writes_value_to_correct_cell(self, tmp_path):
        filepath = self._create_test_excel(tmp_path, [
            ["I2609", "2026-04-10", 650.0, 720.0, "N", None],
            ["I2605", "2026-03-15", 680.0, 710.0, "Y", None],
        ])
        save_captured_basis(filepath, row_index=3, value=25.0)

        # Verify written value
        positions = load_positions(filepath)
        sold_pos = [p for p in positions if p["sold"]][0]
        assert sold_pos["captured_basis"] == 25.0

    def test_does_not_crash_on_missing_file(self, tmp_path):
        filepath = str(tmp_path / "nonexistent.xlsx")
        # Should not raise
        save_captured_basis(filepath, row_index=2, value=10.0)

    def test_overwrites_existing_value(self, tmp_path):
        filepath = self._create_test_excel(tmp_path, [
            ["I2605", "2026-03-15", 680.0, 710.0, "Y", 15.0],
        ])
        save_captured_basis(filepath, row_index=2, value=25.0)

        positions = load_positions(filepath)
        assert positions[0]["captured_basis"] == 25.0


class TestCalculatePositionReturn:
    def _make_position(self, **overrides):
        defaults = {
            "row_index": 2,
            "contract": "I2609",
            "buy_date": date(2026, 4, 10),
            "buy_price": 650.0,
            "base_price_at_buy": 720.0,
            "sold": False,
            "captured_basis": None,
        }
        defaults.update(overrides)
        return defaults

    def test_calculates_all_metrics(self):
        pos = self._make_position()
        result = calculate_position_return(
            position=pos,
            current_futures_price=665.0,
            current_base_price=730.0,
            reference_date=date(2026, 4, 25),
        )
        assert result["contract"] == "I2609"
        assert result["buy_price"] == 650.0
        assert result["current_price"] == 665.0
        assert result["initial_basis"] == 70.0
        assert result["current_basis"] == 65.0
        assert result["captured_basis"] == 5.0
        assert abs(result["convergence_pct"] - (5 / 70 * 100)) < 0.01
        assert result["holding_days"] == 15
        expected_annualized = 5 / 650 * 365 / 15 * 100
        assert abs(result["annualized_return"] - expected_annualized) < 0.01
        assert result["pnl"] == 15.0

    def test_zero_holding_days_returns_none_annualized(self):
        pos = self._make_position(buy_date=date(2026, 4, 25))
        result = calculate_position_return(
            position=pos,
            current_futures_price=665.0,
            current_base_price=730.0,
            reference_date=date(2026, 4, 25),
        )
        assert result["annualized_return"] is None
        assert result["holding_days"] == 0

    def test_zero_initial_basis_returns_none_convergence(self):
        pos = self._make_position(buy_price=720.0, base_price_at_buy=720.0)
        result = calculate_position_return(
            position=pos,
            current_futures_price=720.0,
            current_base_price=720.0,
            reference_date=date(2026, 4, 25),
        )
        assert result["convergence_pct"] is None

    def test_negative_basis_still_calculates(self):
        pos = self._make_position(buy_price=750.0, base_price_at_buy=720.0)
        result = calculate_position_return(
            position=pos,
            current_futures_price=740.0,
            current_base_price=730.0,
            reference_date=date(2026, 4, 25),
        )
        assert result["initial_basis"] == -30.0
        assert result["current_basis"] == -10.0
        assert result["captured_basis"] == -20.0


class TestBuildPositionSummary:
    def test_formats_text_for_wechat(self):
        returns = [
            {
                "contract": "I2609",
                "buy_price": 650.0,
                "current_price": 665.0,
                "captured_basis": 5.0,
                "convergence_pct": 7.14,
                "annualized_return": 18.72,
                "pnl": 15.0,
                "sold": False,
            },
            {
                "contract": "I2612",
                "buy_price": 620.0,
                "current_price": 635.0,
                "captured_basis": 10.0,
                "convergence_pct": 66.67,
                "annualized_return": 98.17,
                "pnl": 15.0,
                "sold": False,
            },
        ]
        result = build_position_summary(returns, total_count=3)
        assert "I2609" in result
        assert "I2612" in result
        assert "650.0" in result
        assert "持仓数: 3" in result
        assert "未平仓: 2" in result

    def test_empty_returns_empty_string(self):
        result = build_position_summary([], total_count=0)
        assert result == ""

    def test_hides_sold_positions_in_detail(self):
        returns = [
            {
                "contract": "I2609",
                "buy_price": 650.0,
                "current_price": 665.0,
                "captured_basis": 5.0,
                "convergence_pct": 7.14,
                "annualized_return": 18.72,
                "pnl": 15.0,
                "sold": False,
            },
            {
                "contract": "I2605",
                "buy_price": 680.0,
                "current_price": 700.0,
                "captured_basis": 20.0,
                "convergence_pct": 100.0,
                "annualized_return": 50.0,
                "pnl": 20.0,
                "sold": True,
            },
        ]
        result = build_position_summary(returns, total_count=2)
        assert "I2609" in result
        assert "I2605" not in result
        assert "持仓数: 2" in result
        assert "未平仓: 1" in result


class TestComputePositionReturns:
    def test_filters_unsold_and_matches_prices(self, tmp_path):
        positions = [
            {"row_index": 2, "contract": "I2609", "buy_date": date(2026, 4, 10),
             "buy_price": 650.0, "base_price_at_buy": 720.0, "sold": False, "captured_basis": None},
            {"row_index": 3, "contract": "I2605", "buy_date": date(2026, 3, 15),
             "buy_price": 680.0, "base_price_at_buy": 710.0, "sold": True, "captured_basis": None},
            {"row_index": 4, "contract": "I2612", "buy_date": date(2026, 4, 12),
             "buy_price": 630.0, "base_price_at_buy": 720.0, "sold": False, "captured_basis": None},
        ]
        price_map = {"I2609": 665.0, "I2612": 635.0, "I2605": 700.0}
        near_price = 730.0

        returns = compute_position_returns(positions, price_map, near_price)
        assert len(returns) == 3  # 2 unsold + 1 sold without captured_basis
        unsold = [r for r in returns if not r["sold"]]
        assert len(unsold) == 2
        assert returns[0]["contract"] == "I2609"
        assert returns[1]["contract"] == "I2605"
        assert returns[2]["contract"] == "I2612"

    def test_skips_positions_with_no_matching_price(self):
        positions = [
            {"row_index": 2, "contract": "I2701", "buy_date": date(2026, 4, 10),
             "buy_price": 600.0, "base_price_at_buy": 700.0, "sold": False, "captured_basis": None},
        ]
        price_map = {"I2609": 665.0}
        near_price = 730.0

        returns = compute_position_returns(positions, price_map, near_price)
        assert len(returns) == 0

    def test_skips_sold_with_frozen_captured_basis(self):
        positions = [
            {"row_index": 2, "contract": "I2605", "buy_date": date(2026, 3, 15),
             "buy_price": 680.0, "base_price_at_buy": 710.0, "sold": True, "captured_basis": 25.0},
        ]
        price_map = {"I2605": 700.0}
        near_price = 730.0

        returns = compute_position_returns(positions, price_map, near_price)
        assert len(returns) == 0
